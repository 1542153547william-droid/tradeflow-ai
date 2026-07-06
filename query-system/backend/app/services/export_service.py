"""导出服务：把 SearchResult（分层结构）扁平化导出为 Excel / CSV。"""
from __future__ import annotations

import io
from typing import Callable, List, Literal, Tuple

from ..models import Product, SearchResult


def _rank(p: Product):
    sc = p.search_context
    if sc.organic_rank is not None:
        return f"自然#{sc.organic_rank}"
    if sc.sponsored_rank is not None:
        return f"广告#{sc.sponsored_rank}"
    return ""


def _coupon(p: Product):
    return p.pricing.coupon.text if p.pricing.coupon else ""


# (列标题, 取值函数) —— 把分层字段扁平成一行
_COLUMNS: List[Tuple[str, Callable[[Product], object]]] = [
    ("平台", lambda p: p.base_info.platform),
    ("排名", _rank),
    ("商品ID", lambda p: p.base_info.product_id),
    ("标题", lambda p: p.base_info.title),
    ("品牌", lambda p: p.base_info.brand),
    ("现价", lambda p: p.pricing.price),
    ("货币", lambda p: p.pricing.currency),
    ("原价", lambda p: p.pricing.list_price),
    ("折扣%", lambda p: p.pricing.discount_pct),
    ("优惠券", _coupon),
    ("评分", lambda p: p.base_info.rating),
    ("评论数", lambda p: p.base_info.review_count),
    ("销量排名", lambda p: p.base_info.rank),
    ("排名类目", lambda p: p.base_info.rank_category),
    ("徽章", lambda p: " / ".join(p.base_info.badges)),
    ("卖家", lambda p: p.logistics.seller),
    ("发货", lambda p: p.logistics.fulfillment),
    ("尺寸", lambda p: p.logistics.dimensions),
    ("重量", lambda p: p.logistics.weight),
    ("图片数", lambda p: p.content.image_count),
    ("有视频", lambda p: "是" if p.content.has_video else ""),
    ("卖点数", lambda p: len(p.content.bullet_points)),
    ("快速配送", lambda p: p.pricing.fast_shipping),
    ("链接", lambda p: p.base_info.product_url),
]


def _rows(result: SearchResult):
    for p in result.products:
        yield {label: fn(p) for label, fn in _COLUMNS}


def to_csv_bytes(result: SearchResult) -> bytes:
    import csv

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=[label for label, _ in _COLUMNS])
    writer.writeheader()
    for row in _rows(result):
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")  # BOM 便于 Excel 打开中文


def to_xlsx_bytes(result: SearchResult) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "TOP产品"

    header_fill = PatternFill("solid", fgColor="FF9900")  # Amazon 橙
    header_font = Font(bold=True, color="FFFFFF")
    headers = [label for label, _ in _COLUMNS]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows = list(_rows(result))
    for row in rows:
        ws.append([row[label] for label, _ in _COLUMNS])

    # 列宽自适应（粗略）
    for col, (label, _) in enumerate(_COLUMNS, 1):
        max_len = max([len(str(label))] + [len(str(r[label] or "")) for r in rows] or [10])
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"

    # 第二张表：评论分析
    if result.review_analysis:
        ra = result.review_analysis
        ws2 = wb.create_sheet("评论分析")
        ws2.append(["指标", "值"])
        ws2.append(["评论总数", ra.total_reviews])
        ws2.append(["情感得分(0-1)", ra.sentiment_score])
        ws2.append(["正面占比", ra.positive_ratio])
        ws2.append(["中性占比", ra.neutral_ratio])
        ws2.append(["负面占比", ra.negative_ratio])
        ws2.append([])
        ws2.append(["Top 关键词", "权重"])
        for kw in ra.top_keywords:
            ws2.append([kw.keyword, kw.weight])
        for cell in ws2["1:1"]:
            cell.font = Font(bold=True)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export(result: SearchResult, fmt: Literal["xlsx", "csv"]) -> Tuple[bytes, str, str]:
    """返回 (bytes, media_type, filename)。"""
    safe_kw = "".join(c for c in result.keyword if c.isalnum() or c in " -_")[:40].strip() or "result"
    if fmt == "csv":
        return to_csv_bytes(result), "text/csv", f"amazon_{safe_kw}.csv"
    return (
        to_xlsx_bytes(result),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        f"amazon_{safe_kw}.xlsx",
    )
