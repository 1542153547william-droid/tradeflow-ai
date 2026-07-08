"""数据加载层（任务 0.3）—— 所有智能体读 `data/<agent>/` 的唯一入口。

设计目标（配套 data/README.md 的 schema 约定）：
- 一个函数 `load_data(agent, name)` 读表格（CSV/Excel）返回 `list[dict]`，读
  文档（md）返回原始字符串。工具和 prompt 都走这里，**不在业务代码里硬编码禁词表
  等内容**。
- 缺文件 → 返回空（`[]` / `""`），不崩，支持"代码先行、真实数据后补"。
- 声明了 `required_columns` 却缺列 → **显式报错**（G2：不静默出错）。

Excel 依赖 openpyxl（惰性导入）：不装也能 import 本模块、读 CSV/md；只有真的去读
`.xlsx` 才需要它。
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Union

# data/ 在项目根目录；本文件位于 tradeflow/ 下，故上溯一级再进 data。
DATA_ROOT = Path(__file__).resolve().parents[1] / "data"

# name 未带扩展名时的探测顺序：表格优先，其次文档。
_TABLE_EXTS = (".csv", ".xlsx", ".xlsm")
_DOC_EXTS = (".md", ".txt")

Row = Dict[str, str]
Loaded = Union[List[Row], str]


class DataError(RuntimeError):
    """数据存在但不合规（如缺必填列）时抛出——区别于"文件不存在"的静默回退。"""


def data_path(agent: str, name: str) -> Optional[Path]:
    """解析 `data/<agent>/<name>` 的真实路径；带扩展名直接用，否则按约定探测。
    找不到返回 None（交给 load_data 决定回退还是报错）。"""
    base = DATA_ROOT / agent
    candidate = base / name
    if candidate.suffix:                       # 已带扩展名
        return candidate if candidate.exists() else None
    for ext in (*_TABLE_EXTS, *_DOC_EXTS):     # 未带扩展名，逐个探测
        probe = base / f"{name}{ext}"
        if probe.exists():
            return probe
    return None


def load_data(
    agent: str,
    name: str,
    *,
    required_columns: Optional[Sequence[str]] = None,
    default: Optional[Loaded] = None,
) -> Loaded:
    """读取某智能体的一份数据。

    agent / name：定位到 `data/<agent>/<name>`；name 可省略扩展名（自动探测）。
    required_columns：表格型数据的必填列；缺列抛 `DataError`（G2 显式报错）。
    default：文件不存在时的回退值；未指定时，表格回退 `[]`、文档回退 `""`。
    """
    path = data_path(agent, name)
    if path is None:                           # 缺文件：静默回退，不崩
        if default is not None:
            return default
        return "" if name.endswith(_DOC_EXTS) else []

    ext = path.suffix.lower()
    if ext in _DOC_EXTS:
        return path.read_text(encoding="utf-8")
    if ext == ".csv":
        rows = _read_csv(path)
    elif ext in (".xlsx", ".xlsm"):
        rows = _read_excel(path)
    else:
        raise DataError(f"不支持的数据格式: {path.name}")

    if required_columns:
        _check_columns(path, rows, required_columns)
    return rows


def _read_csv(path: Path) -> List[Row]:
    # utf-8-sig 自动吃掉 Excel 存出的 BOM，避免第一列表头带 ﻿。
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return [
            {(k or "").strip(): (v or "").strip() for k, v in row.items()}
            for row in reader
        ]


def _read_excel(path: Path) -> List[Row]:
    try:
        from openpyxl import load_workbook          # 惰性导入，见模块头说明
    except ImportError as exc:                       # pragma: no cover
        raise DataError(
            f"读取 {path.name} 需要 openpyxl：请 pip install openpyxl。"
        ) from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return []
    out: List[Row] = []
    for raw in rows_iter:
        if raw is None or all(c is None for c in raw):
            continue
        out.append({
            header[i]: ("" if c is None else str(c).strip())
            for i, c in enumerate(raw) if i < len(header) and header[i]
        })
    wb.close()
    return out


def _check_columns(path: Path, rows: List[Row], required: Sequence[str]) -> None:
    present = set(rows[0].keys()) if rows else set()
    missing = [c for c in required if c not in present]
    if missing:
        raise DataError(
            f"{path.name} 缺少必填列 {missing}；现有列: {sorted(present)}"
        )


__all__ = ["load_data", "data_path", "DataError", "DATA_ROOT"]
