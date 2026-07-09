"""#4 广告优化分析智能体的工具集（4.1~4.5）。

替代人工"下报表→筛词→控 ACOS→上否定"的确定性部分；解释与动作组合交给模型。

- `ad_overview()`             —— 4.2 指标统计：整体 + 按广告活动汇总。
- `sku_breakeven_acos()`      —— 盈亏线：结算报告(回款率) + 成本表 → 每 SKU 盈亏平衡 ACOS。
- `classify_search_terms()`   —— 4.3 三分类：好词 / 潜力长尾词 / 垃圾词（+数据不足观察）。
- `suggest_bid_actions()`     —— 4.4 调价建议：加价/降价/否定/观察，幅度出自调词规则表。
- `export_negative_keywords()`—— 4.5 否定词清单：导出可直接上传的 CSV。

数据（见 data/README.md 的 ads schema）：
- 搜索词报告：亚马逊后台原样导出（中文列名，按天×搜索词），`*_真实.*` 不入库。
- 结算报告：付款-日期范围报告交易明细，只取 type=Order 行算回款率。
- 广告活动映射 / 成本表 / 调词规则：阈值与映射全部外置（G3 改表不改代码）。

真实 xlsx 存在维度声明损坏、多月多 sheet 的情况，data_loader 的通用读取覆盖不了，
本模块自带"扫全部 sheet + 按关键列定位表头"的健壮读取。
"""

from __future__ import annotations

import csv
from collections import defaultdict
from datetime import date
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from ..data_loader import DATA_ROOT, DataError, load_data
from .base import tool

ADS_DIR = DATA_ROOT / "ads"

# 报表文件探测顺序：真实数据在本地就用真实的，否则退回仓库里的样例（可测试）。
_AD_REPORTS = ("搜索词报告_真实.xlsx", "搜索词报告_样例.csv")
_SETTLE_REPORTS = ("结算报告_真实.xlsx", "结算报告_样例.csv")
_MAPPING_FILES = ("广告活动映射_真实.csv", "广告活动映射.csv")

# 搜索词报告列名（中文后台导出）。
_COL_CAMP, _COL_TERM = "广告活动名称", "客户搜索词"
_COL_IMP, _COL_CLK, _COL_SPEND = "展示量", "点击量", "花费"
_COL_SALES, _COL_ORDERS = "7天总销售额", "7天总订单数(#)"


def _num(v: Any) -> float:
    """报表数值兜底：空串/None/'-'（如无销售时的 ACOS 列）一律按 0。"""
    if v is None or v == "" or v == "-":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _pick(candidates: Tuple[str, ...]) -> Optional[str]:
    for name in candidates:
        if (ADS_DIR / name).exists():
            return name
    return None


def _read_sheets(name: str, key_columns: Tuple[str, ...]) -> List[Dict[str, Any]]:
    """读 data/ads/ 下的一份报表，返回 list[dict]。

    csv 走 data_loader；xlsx 扫**全部 sheet**，在每个 sheet 的前几行里找同时含
    key_columns 的行当表头（真实结算报告每月一个 sheet、且混有透视表 sheet——
    透视表列名不同会被自然跳过；真实搜索词报告维度声明损坏，read_only 读不全，
    这里用完整模式解析）。
    """
    path = ADS_DIR / name
    if not path.exists():
        return []
    if path.suffix.lower() == ".csv":
        rows = load_data("ads", name)
        if not rows or not all(c in rows[0] for c in key_columns):
            return []
        return rows
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise DataError(f"读取 {name} 需要 openpyxl：请 pip install openpyxl。") from exc

    out: List[Dict[str, Any]] = []
    wb = load_workbook(path, data_only=True)   # 非 read_only：容忍损坏的维度声明
    for ws in wb.worksheets:
        header: List[str] = []
        for row in ws.iter_rows(values_only=True):
            if not header:
                cells = [str(c).strip() if c is not None else "" for c in row]
                if all(k in cells for k in key_columns):
                    header = cells
                continue
            if all(c is None for c in row):
                continue
            out.append({header[i]: c for i, c in enumerate(row)
                        if i < len(header) and header[i]})
    wb.close()
    return out


@lru_cache(maxsize=None)
def _rules() -> Dict[str, float]:
    rows = load_data("ads", "调词规则.csv", required_columns=["参数", "值"])
    return {r["参数"]: _num(r["值"]) for r in rows}


def _rule(key: str, fallback: float) -> float:
    return _rules().get(key, fallback)


@lru_cache(maxsize=None)
def _campaign_to_sku() -> Dict[str, str]:
    name = _pick(_MAPPING_FILES)
    if not name:
        return {}
    rows = load_data("ads", name, required_columns=["广告活动名称", "SKU"])
    return {r["广告活动名称"]: r["SKU"] for r in rows if r.get("SKU")}


@lru_cache(maxsize=None)
def _sku_stats() -> Dict[str, Dict[str, float]]:
    """结算报告 type=Order 行 → 每 SKU {销量, 销售额, 实收, 回款率, 客单价}。"""
    name = _pick(_SETTLE_REPORTS)
    if not name:
        return {}
    acc: Dict[str, List[float]] = defaultdict(lambda: [0.0, 0.0, 0.0])
    for r in _read_sheets(name, ("type", "sku", "product sales", "total")):
        if str(r.get("type", "")).strip() != "Order":
            continue
        sku = str(r.get("sku", "")).strip()
        if not sku:
            continue
        acc[sku][0] += _num(r.get("quantity"))
        acc[sku][1] += _num(r.get("product sales"))
        acc[sku][2] += _num(r.get("total"))
    out: Dict[str, Dict[str, float]] = {}
    for sku, (qty, sales, net) in acc.items():
        if sales <= 0:
            continue
        out[sku] = {"销量": qty, "销售额": round(sales, 2), "实收": round(net, 2),
                    "回款率": round(net / sales, 4),
                    "客单价": round(sales / qty, 2) if qty else 0.0}
    return out


@lru_cache(maxsize=None)
def _cost_table() -> Dict[str, float]:
    rows = load_data("ads", "成本表.csv", default=[])
    out = {}
    for r in rows:
        sku = (r.get("SKU") or "").strip()
        if sku:
            out[sku] = _num(r.get("采购成本")) + _num(r.get("头程运费"))
    return out


def _breakeven(sku: str) -> Tuple[float, str]:
    """某 SKU 的盈亏平衡 ACOS 及其口径说明。

    有结算+成本：回款率 − 货成本/客单价（真实盈亏线）。
    有结算无成本：回款率 × 0.5（保守估计货成占实收一半）。
    什么都没有：调词规则表的「默认盈亏ACOS」。
    """
    stats = _sku_stats().get(sku)
    if stats:
        unit_cost = _cost_table().get(sku)
        if unit_cost is not None and stats["客单价"]:
            be = stats["回款率"] - unit_cost / stats["客单价"]
            return round(max(be, 0.0), 4), "真实盈亏线（结算回款率−货成本占比）"
        return round(stats["回款率"] * 0.5, 4), "估算盈亏线（回款率×0.5，成本表缺该SKU）"
    return _rule("默认盈亏ACOS", 0.30), "默认盈亏线（无结算/映射数据）"


def _campaign_breakeven(campaign: str) -> Tuple[float, str, str]:
    sku = _campaign_to_sku().get(campaign, "")
    be, basis = _breakeven(sku) if sku else (_rule("默认盈亏ACOS", 0.30),
                                             "默认盈亏线（广告活动未映射到SKU）")
    return be, basis, sku


def _aggregated_terms(report: str) -> List[Dict[str, Any]]:
    """按 (广告活动, 搜索词) 跨天聚合，并算派生指标。"""
    rows = _read_sheets(report, (_COL_CAMP, _COL_TERM, _COL_CLK, _COL_SPEND))
    acc: Dict[Tuple[str, str], List[float]] = defaultdict(lambda: [0.0] * 5)
    for r in rows:
        key = (str(r.get(_COL_CAMP, "")).strip(), str(r.get(_COL_TERM, "")).strip())
        a = acc[key]
        a[0] += _num(r.get(_COL_IMP)); a[1] += _num(r.get(_COL_CLK))
        a[2] += _num(r.get(_COL_SPEND)); a[3] += _num(r.get(_COL_SALES))
        a[4] += _num(r.get(_COL_ORDERS))
    out = []
    for (camp, term), (imp, clk, spend, sales, orders) in acc.items():
        out.append({
            "广告活动": camp, "搜索词": term,
            "展示": int(imp), "点击": int(clk),
            "花费": round(spend, 2), "销售额": round(sales, 2), "订单": int(orders),
            "ACOS": round(spend / sales, 4) if sales else None,
            "CTR": round(clk / imp, 4) if imp else 0.0,
            "转化率": round(orders / clk, 4) if clk else 0.0,
            "CPC": round(spend / clk, 2) if clk else 0.0,
        })
    return out


def _classify(t: Dict[str, Any], breakeven: float) -> Tuple[str, str]:
    """单个聚合后搜索词的三分类（阈值全部来自调词规则表）。"""
    clicks, orders, acos = t["点击"], t["订单"], t["ACOS"]
    if orders == 0 and clicks >= _rule("垃圾词_零转化最低点击", 12):
        return "垃圾词", f"点击{clicks}零转化（阈值{_rule('垃圾词_零转化最低点击', 12):.0f}）"
    if acos is not None and acos >= breakeven * _rule("垃圾词_ACOS倍数", 1.5):
        return "垃圾词", f"ACOS {acos:.0%} ≥ 盈亏线{breakeven:.0%}×{_rule('垃圾词_ACOS倍数', 1.5)}"
    if (orders >= _rule("好词_最低订单数", 2)
            and acos is not None and acos <= breakeven * _rule("好词_ACOS倍数", 0.8)):
        return "好词", f"订单{orders}且 ACOS {acos:.0%} ≤ 盈亏线{breakeven:.0%}×{_rule('好词_ACOS倍数', 0.8)}"
    if orders >= 1 and clicks >= _rule("潜力词_最低点击", 3):
        return "潜力长尾词", f"已出单{orders}，数据量小待验证"
    return "观察", "数据不足，暂不动作"


@tool
def ad_overview(report: str = "") -> Dict[str, Any]:
    """4.2 指标统计：整体 + 按广告活动汇总（花费/销售/ACOS/点击/订单）。

    report 留空自动选报表（本地真实数据优先，否则样例）。"""
    report = report or _pick(_AD_REPORTS) or ""
    terms = _aggregated_terms(report)
    if not terms:
        return {"error": f"读不到搜索词报告（data/ads/ 下未找到 {_AD_REPORTS}）"}
    camps: Dict[str, List[float]] = defaultdict(lambda: [0.0] * 4)
    for t in terms:
        c = camps[t["广告活动"]]
        c[0] += t["花费"]; c[1] += t["销售额"]; c[2] += t["点击"]; c[3] += t["订单"]
    by_campaign = []
    for name, (spend, sales, clk, orders) in sorted(camps.items(), key=lambda x: -x[1][0]):
        be, basis, sku = _campaign_breakeven(name)
        by_campaign.append({
            "广告活动": name, "SKU": sku or "未映射",
            "花费": round(spend, 2), "销售额": round(sales, 2),
            "ACOS": round(spend / sales, 4) if sales else None,
            "点击": int(clk), "订单": int(orders),
            "盈亏平衡ACOS": be, "盈亏线口径": basis,
        })
    total_spend = sum(c["花费"] for c in by_campaign)
    total_sales = sum(c["销售额"] for c in by_campaign)
    return {"报表": report, "搜索词数": len(terms),
            "总花费": round(total_spend, 2), "总销售额": round(total_sales, 2),
            "总ACOS": round(total_spend / total_sales, 4) if total_sales else None,
            "按广告活动": by_campaign}


@tool
def sku_breakeven_acos() -> Dict[str, Any]:
    """每个 SKU 的回款率与盈亏平衡 ACOS（结算报告 + 成本表联动）。"""
    stats = _sku_stats()
    if not stats:
        return {"error": f"读不到结算报告（data/ads/ 下未找到 {_SETTLE_REPORTS}）"}
    out = []
    for sku, s in sorted(stats.items(), key=lambda x: -x[1]["销售额"]):
        be, basis = _breakeven(sku)
        out.append({**{"SKU": sku}, **s, "盈亏平衡ACOS": be, "口径": basis})
    return {"SKU数": len(out), "明细": out,
            "说明": "广告 ACOS 高于该 SKU 盈亏平衡 ACOS 即在亏钱"}


@tool
def classify_search_terms(report: str = "", campaign: str = "",
                          top_n: int = 30) -> Dict[str, Any]:
    """4.3 搜索词三分类：好词 / 潜力长尾词 / 垃圾词 / 观察（按各活动盈亏线判定）。

    campaign 非空则只看该广告活动（包含匹配）；每类按花费倒序取前 top_n。"""
    report = report or _pick(_AD_REPORTS) or ""
    terms = _aggregated_terms(report)
    if not terms:
        return {"error": f"读不到搜索词报告（data/ads/ 下未找到 {_AD_REPORTS}）"}
    low = campaign.lower().strip()
    buckets: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for t in terms:
        if low and low not in t["广告活动"].lower():
            continue
        be, basis, _sku = _campaign_breakeven(t["广告活动"])
        cls, reason = _classify(t, be)
        buckets[cls].append({**t, "判定依据": reason, "盈亏线": be})
    for cls in buckets:
        buckets[cls].sort(key=lambda x: -x["花费"])
    return {"报表": report, "筛选活动": campaign or "全部",
            "统计": {k: len(v) for k, v in buckets.items()},
            "分类": {k: v[:top_n] for k, v in buckets.items()}}


@tool
def suggest_bid_actions(report: str = "", campaign: str = "",
                        top_n: int = 30) -> Dict[str, Any]:
    """4.4 调价/否定建议：每个搜索词给出 动作+幅度+理由（幅度出自调词规则表）。

    无当前竞价数据，加减价基于实际 CPC 估算（上传前人工过目）。"""
    result = classify_search_terms.func(report, campaign, top_n=top_n)
    if "error" in result:
        return result
    up, down = _rule("建议_加价幅度", 0.10), _rule("建议_降价幅度", 0.15)
    actions = []
    for cls, rows in result["分类"].items():
        for t in rows:
            if cls == "好词":
                act, new_bid = f"加价 {up:.0%}", round(t["CPC"] * (1 + up), 2)
            elif cls == "垃圾词":
                act, new_bid = "加否定（精准）", None
            elif cls == "潜力长尾词":
                act, new_bid = "维持竞价，建议转手动精准拓词", None
            else:
                acos, be = t["ACOS"], t["盈亏线"]
                if acos is not None and acos > be:
                    act, new_bid = f"降价 {down:.0%}", round(t["CPC"] * (1 - down), 2)
                else:
                    act, new_bid = "观察", None
            actions.append({"搜索词": t["搜索词"], "广告活动": t["广告活动"],
                            "分类": cls, "动作": act, "参考CPC": t["CPC"],
                            "建议竞价": new_bid, "理由": t["判定依据"]})
    order = {"垃圾词": 0, "好词": 1, "潜力长尾词": 2, "观察": 3}
    actions.sort(key=lambda a: order.get(a["分类"], 9))
    return {"报表": result["报表"], "筛选活动": result["筛选活动"],
            "动作数": len(actions), "动作清单": actions}


@tool
def export_negative_keywords(report: str = "", campaign: str = "") -> Dict[str, Any]:
    """4.5 导出否定词清单 CSV（垃圾词 → 否定精准），可直接上传亚马逊后台。

    写入 data/ads/输出/否定词清单_<日期>.csv，返回路径与条数。"""
    result = classify_search_terms.func(report, campaign, top_n=10_000)
    if "error" in result:
        return result
    rows = result["分类"].get("垃圾词", [])
    if not rows:
        return {"条数": 0, "说明": "本次没有达到否定标准的搜索词"}
    out_dir = ADS_DIR / "输出"
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"否定词清单_{date.today():%Y%m%d}.csv"
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Campaign Name", "Ad Group Name", "Customer Search Term",
                    "Match Type", "花费", "点击", "判定依据"])
        for t in rows:
            w.writerow([t["广告活动"], t["广告活动"], t["搜索词"],
                        "Negative Exact", t["花费"], t["点击"], t["判定依据"]])
    return {"条数": len(rows), "文件": str(path),
            "预计月省": round(sum(t["花费"] for t in rows), 2),
            "说明": "按否定精准导出；上传前请人工过目"}


def reset_caches() -> None:
    """测试用：清掉模块级缓存（数据文件变化后重读）。"""
    for fn in (_rules, _campaign_to_sku, _sku_stats, _cost_table):
        fn.cache_clear()


ADS_TOOLS = [ad_overview, sku_breakeven_acos, classify_search_terms,
             suggest_bid_actions, export_negative_keywords]
