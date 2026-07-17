"""Excel/CSV preview, field mapping, and durable import."""

from __future__ import annotations

import csv
import io
import json
import re
import uuid
from collections.abc import Iterable, Iterator
from typing import Any

from web.database import audit, connect

HEADER_SCAN_ROWS = 30
PREVIEW_ROWS = 8
INSERT_CHUNK_SIZE = 1000

ALIASES = {
    "sku": {"sku", "seller sku", "商家sku", "卖家sku"},
    "asin": {"asin", "商品asin"},
    "campaign": {"campaign name", "广告活动名称", "广告活动"},
    "search_term": {"customer search term", "客户搜索词", "搜索词"},
    "impressions": {"impressions", "展示量", "曝光量"},
    "clicks": {"clicks", "点击量"},
    "spend": {"spend", "花费", "广告花费"},
    "sales": {"7 day total sales", "7天总销售额", "销售额", "product sales"},
    "orders": {"7 day total orders", "7天总订单数(#)", "订单数", "quantity"},
    "quantity": {"quantity", "qty", "数量", "购买数量"},
    "price": {"price", "价格", "售价"},
    "stock": {"stock", "库存", "库存数量"},
    "order_id": {"amazon-order-id", "order id", "订单号"},
    "buyer": {"buyer", "buyer name", "customer", "customer name", "recipient name", "ship to name", "买家", "买家姓名", "收件人"},
    "purchase_date": {"purchase date", "order date", "date", "下单时间", "购买日期", "下单日期"},
    "title": {"title", "product title", "商品标题", "标题"},
    "rating": {"rating", "star rating", "评分", "星级"},
    "review_count": {"review count", "reviews", "评论数", "评价数"},
    "brand": {"brand", "品牌"},
}


def _norm(value: Any) -> str:
    return re.sub(r"[\s_\-]+", " ", str(value or "").strip().lower())


def suggest_mapping(columns: list[str]) -> dict[str, str]:
    out: dict[str, str] = {}
    for col in columns:
        n = _norm(col)
        for target, aliases in ALIASES.items():
            if n in {_norm(a) for a in aliases}:
                out[col] = target
                break
    return out


def detect_report_type(mapping: dict[str, str]) -> str:
    fields = set(mapping.values())
    if {"campaign", "search_term", "clicks", "spend"} <= fields:
        return "ads_search_terms"
    if "order_id" in fields or {"sku", "orders", "sales"} <= fields:
        return "orders"
    if {"sku", "stock"} <= fields:
        return "inventory"
    if "asin" in fields:
        return "competitors"
    return "unknown"


def _is_nonempty(row: list[Any]) -> bool:
    return any(v not in (None, "") for v in row)


def _header_score(row: list[Any], known: set[str]) -> int:
    normalized = [_norm(c) for c in row if _norm(c)]
    recognized = sum(c in known for c in normalized)
    # Business-field matches dominate; fuller rows win exact ties.
    return recognized * 100 + len(normalized)


def _iter_csv_rows(content: bytes) -> Iterator[list[Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    yield from csv.reader(io.StringIO(text))


def _iter_xlsx_rows(content: bytes, *, read_only: bool) -> Iterator[list[Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=read_only)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                yield list(row)
    finally:
        wb.close()


def _iter_upload_rows(filename: str, content: bytes, *, read_only: bool = True) -> Iterator[list[Any]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        yield from _iter_csv_rows(content)
    elif lower.endswith((".xlsx", ".xlsm")):
        yield from _iter_xlsx_rows(content, read_only=read_only)
    else:
        raise ValueError("仅支持 .xlsx、.xlsm 和 .csv")


def _parse_rows(rows_iter: Iterable[list[Any]], *, data_limit: int | None = None) -> tuple[list[str], list[dict[str, Any]], int]:
    known = {_norm(a) for aliases in ALIASES.values() for a in aliases}
    scan: list[list[Any]] = []
    iterator = iter(rows_iter)
    for row in iterator:
        if not _is_nonempty(row):
            continue
        scan.append(row)
        if len(scan) >= HEADER_SCAN_ROWS:
            break
    if not scan:
        raise ValueError("文件中没有可读取的数据")

    scored = [(_header_score(row, known), -i, i) for i, row in enumerate(scan)]
    _, _, header_idx = max(scored)
    columns = [str(v or f"column_{i+1}").strip() for i, v in enumerate(scan[header_idx])]
    data = []
    row_count = 0

    def consume(row: list[Any]) -> None:
        nonlocal row_count
        item = {columns[i]: row[i] for i in range(min(len(columns), len(row))) if row[i] not in (None, "")}
        if item:
            row_count += 1
            if data_limit is None or len(data) < data_limit:
                data.append(item)

    for row in scan[header_idx + 1:]:
        consume(row)
    for row in iterator:
        if _is_nonempty(row):
            consume(row)
    return columns, data, row_count


def _parse_upload(filename: str, content: bytes, *, data_limit: int | None = None) -> tuple[list[str], list[dict[str, Any]], int]:
    lower = filename.lower()
    columns, rows, row_count = _parse_rows(_iter_upload_rows(filename, content, read_only=True), data_limit=data_limit)
    # Amazon exports occasionally declare a broken worksheet dimension
    # (for example A1:A1 although thousands of cells exist). read_only mode
    # trusts that declaration. Fast-path normal files, and fall back only when
    # the parsed shape looks suspicious.
    if lower.endswith((".xlsx", ".xlsm")) and (len(columns) <= 1 or row_count == 0):
        columns, rows, row_count = _parse_rows(_iter_upload_rows(filename, content, read_only=False), data_limit=data_limit)
    return columns, rows, row_count


def parse_upload(filename: str, content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    columns, rows, _ = _parse_upload(filename, content)
    return columns, rows


def parse_upload_preview(filename: str, content: bytes) -> dict[str, Any]:
    columns, preview, row_count = _parse_upload(filename, content, data_limit=PREVIEW_ROWS)
    mapping = suggest_mapping(columns)
    return {"columns": columns, "mapping": mapping, "preview": preview, "row_count": row_count}


def save_import(user_id: str, store_id: str, filename: str, columns: list[str],
                rows: list[dict[str, Any]], mapping: dict[str, str]) -> dict[str, Any]:
    batch_id = f"imp_{uuid.uuid4().hex[:12]}"
    report_type = detect_report_type(mapping)
    row_count = len(rows)
    with connect() as db:
        db.execute(
            "INSERT INTO import_batches(id,user_id,store_id,filename,report_type,status,row_count,columns_json,mapping_json) VALUES(?,?,?,?,?,'completed',?,?,?)",
            (batch_id, user_id, store_id, filename, report_type, row_count,
             json.dumps(columns, ensure_ascii=False), json.dumps(mapping, ensure_ascii=False)),
        )
        chunk = []
        for row in rows:
            normalized = {mapping.get(k, k): v for k, v in row.items()}
            chunk.append((batch_id, user_id, store_id, json.dumps(normalized, ensure_ascii=False, default=str)))
            if len(chunk) >= INSERT_CHUNK_SIZE:
                db.executemany(
                    "INSERT INTO imported_rows(batch_id,user_id,store_id,row_json) VALUES(?,?,?,?)",
                    chunk,
                )
                chunk.clear()
        if chunk:
            db.executemany(
                "INSERT INTO imported_rows(batch_id,user_id,store_id,row_json) VALUES(?,?,?,?)",
                chunk,
            )
    audit(user_id, store_id, "import.completed", {"batch_id": batch_id, "rows": row_count})
    return {"id": batch_id, "filename": filename, "report_type": report_type,
            "row_count": row_count, "status": "completed"}


def list_imports(user_id: str, store_id: str) -> list[dict[str, Any]]:
    with connect() as db:
        rows = db.execute(
            "SELECT id,filename,report_type,status,row_count,created_at FROM import_batches WHERE user_id=? AND store_id=? ORDER BY created_at DESC",
            (user_id, store_id),
        ).fetchall()
    return [dict(r) for r in rows]


def _latest_imported_rows(user_id: str, store_id: str, report_type: str) -> list[dict[str, Any]]:
    with connect() as db:
        batch = db.execute(
            "SELECT id FROM import_batches WHERE user_id=? AND store_id=? AND report_type=? "
            "AND status='completed' AND row_count>0 ORDER BY created_at DESC, id DESC LIMIT 1",
            (user_id, store_id, report_type),
        ).fetchone()
        if not batch:
            return []
        rows = db.execute(
            "SELECT row_json FROM imported_rows WHERE user_id=? AND store_id=? AND batch_id=?",
            (user_id, store_id, batch["id"]),
        ).fetchall()
    return [json.loads(r[0]) for r in rows]


def ads_overview(user_id: str, store_id: str) -> dict[str, Any]:
    data = _latest_imported_rows(user_id, store_id, "ads_search_terms")
    if not data:
        return {"items": [], "error": "请先导入真实广告搜索词报表"}
    campaigns: dict[str, dict[str, Any]] = {}
    for row in data:
        name = str(row.get("campaign") or "未命名广告活动")
        c = campaigns.setdefault(name, {"campaign": name, "impressions": 0.0, "clicks": 0.0,
                                        "spend": 0.0, "sales": 0.0, "orders": 0.0})
        for key in ("impressions", "clicks", "spend", "sales", "orders"):
            try:
                c[key] += float(str(row.get(key, 0)).replace(",", "") or 0)
            except ValueError:
                pass
    items = []
    for c in campaigns.values():
        c["acos"] = round(c["spend"] / c["sales"], 4) if c["sales"] else None
        c["ctr"] = round(c["clicks"] / c["impressions"], 4) if c["impressions"] else 0
        c["conversion_rate"] = round(c["orders"] / c["clicks"], 4) if c["clicks"] else 0
        c["score"] = _ad_health_score(c)
        c["severity"] = _ad_severity(c)
        c["recommendations"] = _ad_recommendations(c)
        items.append(c)
    return {"items": sorted(items, key=lambda x: -x["spend"]), "source": "imported_report",
            "row_count": len(data)}


def _ad_health_score(item: dict[str, Any]) -> int:
    score = 90
    acos = item.get("acos")
    ctr = item.get("ctr") or 0
    cvr = item.get("conversion_rate") or 0
    if acos is None and item.get("spend", 0) >= 5:
        score -= 35
    elif acos is not None:
        if acos >= 0.6:
            score -= 35
        elif acos >= 0.4:
            score -= 22
        elif acos >= 0.25:
            score -= 10
    if ctr < 0.003 and item.get("impressions", 0) >= 1000:
        score -= 12
    if cvr < 0.03 and item.get("clicks", 0) >= 20:
        score -= 14
    return max(0, min(100, round(score)))


def _ad_severity(item: dict[str, Any]) -> str:
    score = item.get("score", 0)
    if score < 55:
        return "crit"
    if score < 78:
        return "warn"
    return "good"


def _ad_recommendations(item: dict[str, Any]) -> list[dict[str, str]]:
    recs: list[dict[str, str]] = []
    spend = float(item.get("spend") or 0)
    sales = float(item.get("sales") or 0)
    orders = float(item.get("orders") or 0)
    clicks = float(item.get("clicks") or 0)
    impressions = float(item.get("impressions") or 0)
    acos = item.get("acos")
    ctr = item.get("ctr") or 0
    cvr = item.get("conversion_rate") or 0
    if spend >= 5 and orders == 0:
        recs.append({"level": "crit", "title": "花费但未出单",
                     "detail": f"花费 {spend:.2f}、点击 {clicks:.0f}、订单 0。建议进入对话查看搜索词，优先否定零转化高花费词。"})
    if acos is not None and acos >= 0.4:
        recs.append({"level": "warn", "title": "ACOS 偏高",
                     "detail": f"ACOS {acos * 100:.1f}%，销售额 {sales:.2f}。建议降竞价、收窄匹配或暂停低效词。"})
    if impressions >= 1000 and ctr < 0.003:
        recs.append({"level": "warn", "title": "点击率偏低",
                     "detail": f"展示 {impressions:.0f}、CTR {ctr * 100:.2f}%。建议优化主图、标题相关性或降低宽泛曝光。"})
    if clicks >= 20 and cvr < 0.03:
        recs.append({"level": "warn", "title": "转化率偏低",
                     "detail": f"点击 {clicks:.0f}、转化率 {cvr * 100:.1f}%。建议检查价格、评价、页面卖点与搜索词意图是否匹配。"})
    if not recs:
        recs.append({"level": "good", "title": "暂无明显广告异常",
                     "detail": "当前活动没有命中高风险规则。建议继续观察，并在对话中结合毛利目标进一步判断预算。"})
    return recs


def customer_overview(user_id: str, store_id: str) -> dict[str, Any]:
    data = _latest_imported_rows(user_id, store_id, "orders")
    if not data:
        return {"items": [], "counts": {"all": 0, "repeat": 0, "review": 0},
                "error": "请先导入真实订单报表"}
    by_buyer: dict[str, int] = {}
    items: list[dict[str, Any]] = []
    for row in data[:300]:
        order_id = str(row.get("order_id") or "").strip()
        sku = str(row.get("sku") or "").strip()
        title = str(row.get("title") or sku or "未命名商品").strip()
        buyer = str(row.get("buyer") or "").strip()
        buyer_key = buyer or order_id or sku or "unknown"
        by_buyer[buyer_key] = by_buyer.get(buyer_key, 0) + 1
        tag = "repeat" if by_buyer[buyer_key] > 1 else "review"
        items.append({
            "name": buyer or "买家信息未导入",
            "prod": title,
            "order": order_id or "订单号未导入",
            "date": str(row.get("purchase_date") or row.get("date") or "时间未导入"),
            "tag": tag,
            "sku": sku,
        })
    counts = {
        "all": len(items),
        "repeat": sum(1 for x in items if x["tag"] == "repeat"),
        "review": sum(1 for x in items if x["tag"] == "review"),
    }
    return {"items": items, "counts": counts, "source": "imported_orders", "row_count": len(data)}


def ads_chat_analysis(user_id: str, store_id: str) -> str | None:
    """Return a concise, actionable ad report analysis for chat.

    The regular agent tools read files from data/ads. Imported user reports live
    in SQLite, so chat needs this bridge to avoid claiming that no report exists.
    """
    data = _latest_imported_rows(user_id, store_id, "ads_search_terms")
    if not data:
        return None

    def num(value: Any) -> float:
        try:
            return float(str(value or 0).replace(",", "").replace("$", "") or 0)
        except ValueError:
            return 0.0

    campaigns: dict[str, dict[str, Any]] = {}
    terms: dict[tuple[str, str], dict[str, Any]] = {}
    for row in data:
        campaign = str(row.get("campaign") or "未命名广告活动")
        term = str(row.get("search_term") or "(空搜索词)").strip() or "(空搜索词)"
        c = campaigns.setdefault(campaign, {"campaign": campaign, "impressions": 0.0, "clicks": 0.0,
                                            "spend": 0.0, "sales": 0.0, "orders": 0.0})
        t = terms.setdefault((campaign, term), {"campaign": campaign, "term": term, "impressions": 0.0,
                                                "clicks": 0.0, "spend": 0.0, "sales": 0.0, "orders": 0.0})
        for key in ("impressions", "clicks", "spend", "sales", "orders"):
            value = num(row.get(key))
            c[key] += value
            t[key] += value

    def enrich(item: dict[str, Any]) -> dict[str, Any]:
        spend, sales, clicks, impressions, orders = (
            item["spend"], item["sales"], item["clicks"], item["impressions"], item["orders"])
        item["acos"] = spend / sales if sales else None
        item["ctr"] = clicks / impressions if impressions else 0.0
        item["cvr"] = orders / clicks if clicks else 0.0
        item["cpc"] = spend / clicks if clicks else 0.0
        return item

    campaign_items = [enrich(dict(v)) for v in campaigns.values()]
    term_items = [enrich(dict(v)) for v in terms.values()]
    high_acos_campaigns = sorted(
        [x for x in campaign_items if x["spend"] >= 10 and (x["acos"] is None or x["acos"] >= 0.4)],
        key=lambda x: x["spend"],
        reverse=True,
    )[:5]
    negatives = sorted(
        [x for x in term_items if x["spend"] >= 5 and x["orders"] == 0],
        key=lambda x: x["spend"],
        reverse=True,
    )[:8]
    reduce_bids = sorted(
        [x for x in term_items if x["spend"] >= 5 and x["sales"] > 0 and x["acos"] and x["acos"] >= 0.4],
        key=lambda x: x["spend"],
        reverse=True,
    )[:6]
    scale_terms = sorted(
        [x for x in term_items if x["orders"] >= 2 and x["acos"] is not None and 0 < x["spend"]
         and x["acos"] <= 0.2],
        key=lambda x: (x["acos"], -x["orders"]),
    )[:8]

    total = enrich({
        "impressions": sum(x["impressions"] for x in campaign_items),
        "clicks": sum(x["clicks"] for x in campaign_items),
        "spend": sum(x["spend"] for x in campaign_items),
        "sales": sum(x["sales"] for x in campaign_items),
        "orders": sum(x["orders"] for x in campaign_items),
    })

    def money(value: float) -> str:
        return f"{value:.2f}"

    def pct(value: float | None) -> str:
        return "无销售" if value is None else f"{value * 100:.1f}%"

    lines = [
        f"已读取你导入的广告搜索词报表，共 {len(data)} 行。按真实入库数据计算：",
        "",
        "整体盘面：",
        f"- 花费 {money(total['spend'])}，销售额 {money(total['sales'])}，订单 {int(total['orders'])}，ACOS {pct(total['acos'])}，转化率 {pct(total['cvr'])}。",
        "",
        "建议否定或重点排查的搜索词：",
    ]
    if negatives:
        for x in negatives:
            lines.append(
                f"- {x['term']}：{x['campaign']}，花费 {money(x['spend'])}，点击 {int(x['clicks'])}，0 单。")
    else:
        lines.append("- 暂无达到阈值的零转化高花费词。")

    lines.extend(["", "建议降价/降预算的广告活动："])
    if high_acos_campaigns:
        for x in high_acos_campaigns:
            lines.append(
                f"- {x['campaign']}：花费 {money(x['spend'])}，销售额 {money(x['sales'])}，订单 {int(x['orders'])}，ACOS {pct(x['acos'])}。")
    else:
        lines.append("- 暂无明显高 ACOS 活动。")

    lines.extend(["", "建议降低竞价的搜索词："])
    if reduce_bids:
        for x in reduce_bids:
            lines.append(
                f"- {x['term']}：{x['campaign']}，花费 {money(x['spend'])}，销售额 {money(x['sales'])}，ACOS {pct(x['acos'])}。")
    else:
        lines.append("- 暂无明显高 ACOS 搜索词。")

    lines.extend(["", "建议加预算/转精准的词："])
    if scale_terms:
        for x in scale_terms:
            lines.append(
                f"- {x['term']}：{x['campaign']}，订单 {int(x['orders'])}，花费 {money(x['spend'])}，销售额 {money(x['sales'])}，ACOS {pct(x['acos'])}。")
    else:
        lines.append("- 目前没有同时满足 2 单以上且 ACOS <= 20% 的放量词。")

    lines.extend([
        "",
        "执行顺序建议：先否定零转化高花费词，再处理高 ACOS 活动，最后把低 ACOS 出单词单独拉到精准广告里放量。",
    ])
    return "\n".join(lines)


def ads_chat_context(user_id: str, store_id: str) -> str | None:
    """Build a compact live data context for the ad agent to reason over."""
    analysis = ads_chat_analysis(user_id, store_id)
    if not analysis:
        return None
    return (
        "以下是系统刚刚从用户已导入的广告搜索词报表中实时聚合出的真实数据摘要。"
        "你必须基于这些数据回答，不要声称没有收到报表，不要使用样例或 Mock 数据。\n\n"
        f"{analysis}\n\n"
        "回答要求：根据用户具体问题选择分析角度；如果用户只笼统要求分析，"
        "请给出最重要的 3-5 条可执行动作，并说明依据、风险和下一步。"
        "金额单位未知，只能沿用数字本身，禁止标 USD、人民币、¥ 或 $。"
        "没有毛利、成本和目标 ACOS 时，禁止计算利润/亏损金额，只能评价花费、销售额和 ACOS。"
        "不要自行把金额翻倍；不要引用上下文里没有提供的规则表参数。"
    )


def competitor_rows(user_id: str, store_id: str, limit: int = 20) -> list[dict[str, Any]]:
    """Return real competitor rows uploaded by this store, newest first."""
    with connect() as db:
        rows = db.execute(
            "SELECT r.row_json FROM imported_rows r JOIN import_batches b ON b.id=r.batch_id "
            "WHERE r.user_id=? AND r.store_id=? AND b.report_type='competitors' "
            "ORDER BY r.id DESC LIMIT ?",
            (user_id, store_id, limit),
        ).fetchall()
    return [json.loads(r[0]) for r in rows]
